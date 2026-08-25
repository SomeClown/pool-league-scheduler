"""
app/main/export.py — Excel workbook generation for season schedules.

Produces a formatted .xlsx workbook with three sheets:
    Schedule — full match listing, colour-coded by home/away/bar columns,
               with alternating row colours per week and bold outline boxes
               around each week's group of matches.
    Teams    — numbered team roster with home bar.
    Bars     — venue list with table counts.

Uses openpyxl for workbook construction. All styling is applied programmatically —
no template files, no magic. If the colours look wrong, the hex values are near
the top of this file and are easy to change. If the layout looks wrong, that's
more work and I'm sorry in advance.

Requires: pip install openpyxl (listed in requirements.txt)
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette — change these if you want a different look.
# All values are hex RGB strings without the leading #.
# ---------------------------------------------------------------------------

BG_TITLE      = "0A2540"   # Very dark navy — sheet title bar
BG_SUBTITLE   = "D6E8F7"   # Pale blue — info band below title
BG_HDR_MAIN   = "1B3A5C"   # Dark navy — Week / Date / Bar column headers
BG_HDR_HOME   = "7B5200"   # Dark amber — Home # / Home Team column headers
BG_HDR_AWAY   = "1A4A7A"   # Dark steel blue — Away # / Away Team column headers
BG_HDR_BAR    = "2D4A2D"   # Dark forest green — Bar column header
BG_ROW_ODD    = "EAF4FC"   # Light sky blue — odd-numbered weeks
BG_ROW_EVEN   = "FFFFFF"   # White — even-numbered weeks
BG_BYE        = "EBEBEB"   # Light grey — bye rows
BG_TEAMS_HDR  = "1B3A5C"   # Same navy used for Teams and Bars sheet headers

FG_WHITE      = "FFFFFF"   # Text on dark backgrounds
FG_DARK       = "111111"   # Standard dark text
FG_MUTED      = "555555"   # Subdued text (bars, bye notes, etc.)


# ---------------------------------------------------------------------------
# Style helpers — small factories to keep the sheet-building code readable.
# These are not complicated. Don't let the number of them fool you.
# ---------------------------------------------------------------------------

def _fill(hex_color):
    """Return a solid PatternFill for the given hex colour."""
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, italic=False, color=FG_DARK, size=11):
    """Return a Font with the specified attributes. Defaults to normal 11pt dark text."""
    return Font(bold=bold, italic=italic, color=color, size=size)


def _align(h="left", v="center", wrap=False):
    """Return an Alignment. Vertical centre by default because top-aligned cells look bad."""
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# Border sides used throughout. _THICK for outlines, _THIN for internal grid lines.
_THICK = Side(border_style="medium", color="222222")
_THIN  = Side(border_style="thin",   color="BBBBBB")
_NONE  = Side(border_style=None)


def _border(top=_NONE, right=_NONE, bottom=_NONE, left=_NONE):
    """Return a Border object with the given sides. Unspecified sides default to _NONE."""
    return Border(top=top, right=right, bottom=bottom, left=left)


def _outline_range(ws, r1, r2, c1, c2):
    """
    Apply a bold (medium) outer border with a thin interior grid to a cell range.

    Each cell in the range gets:
        - Medium (thick) border on edges that touch the outline boundary
        - Thin border on internal edges between cells

    This is the function responsible for the 'bold outline boxes around groupings'
    visual. Called once per week group in the schedule sheet.

    Args:
        ws:         The worksheet to modify.
        r1, r2:     First and last row of the range (1-indexed).
        c1, c2:     First and last column of the range (1-indexed).
    """
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            top    = _THICK if row == r1 else _THIN
            bottom = _THICK if row == r2 else _THIN
            left   = _THICK if col == c1 else _THIN
            right  = _THICK if col == c2 else _THIN
            ws.cell(row=row, column=col).border = _border(top, right, bottom, left)


# ---------------------------------------------------------------------------
# Sheet 1 — Schedule
# Column layout: A=Wk  B=Date  C=Home#  D=Home Team  E=Away#  F=Away Team  G=Bar
# ---------------------------------------------------------------------------

_SCHED_COLS = 7   # Total columns in the schedule sheet.


def _sched_header_cell(ws, row, col, value, bg, align_h="left"):
    """
    Write a single column header cell with the standard header styling.

    Used to build row 3 of the schedule sheet (the column label row).
    Each header cell gets a coloured background, white bold text, and a
    thick border on all four sides. The colour varies by column group
    (home = amber, away = steel blue, bar = green, rest = navy).

    Args:
        ws:      Worksheet to write to.
        row:     Row number (1-indexed).
        col:     Column number (1-indexed).
        value:   Cell text.
        bg:      Background hex colour string.
        align_h: Horizontal alignment ('left' or 'center').
    """
    c           = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(bold=True, color=FG_WHITE, size=10)
    c.alignment = _align(h=align_h)
    c.border    = _border(_THICK, _THICK, _THICK, _THICK)
    return c


def _build_schedule_sheet(ws, season, rounds):
    """
    Populate the Schedule worksheet with season data and formatting.

    Layout:
        Row 1   — Season title (dark navy, merged across all columns)
        Row 2   — Subtitle: dates, frequency, team/round counts (pale blue)
        Row 3   — Column headers (colour-coded by column group)
        Row 4+  — Match data, grouped by week with alternating backgrounds
                  and bold outline borders around each group. Bye rows appear
                  at the end of their week's group in light grey.

    Rows freeze below row 3 so the headers stay visible while scrolling.

    Args:
        ws:      The active (Schedule) worksheet.
        season:  Season model instance.
        rounds:  Ordered dict from _build_rounds() — {round_num: round_data}.
    """
    # ── Row 1 — title ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(_SCHED_COLS)}1")
    c           = ws["A1"]
    c.value     = season.name.upper()
    c.fill      = _fill(BG_TITLE)
    c.font      = _font(bold=True, color=FG_WHITE, size=16)
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 32

    # ── Row 2 — subtitle ─────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(_SCHED_COLS)}2")
    start = season.start_date.strftime("%B %-d, %Y")
    end   = (f" – {season.end_date.strftime('%B %-d, %Y')}" if season.end_date else "")
    sub   = (f"{start}{end}   ·   {season.frequency.title()}   ·   "
             f"{len(season.teams)} teams   ·   {len(rounds)} rounds")
    c           = ws["A2"]
    c.value     = sub
    c.fill      = _fill(BG_SUBTITLE)
    c.font      = _font(color=FG_DARK, size=10)
    c.alignment = _align(h="center")
    ws.row_dimensions[2].height = 18

    # ── Row 3 — column headers ────────────────────────────────────────────────
    # Colour-coded by column group: home columns = amber, away = steel blue,
    # bar = forest green, week/date = navy. This gives you an immediate visual
    # cue about which columns belong to which team without reading the labels.
    hdr_defs = [
        ("Wk",        BG_HDR_MAIN, "center"),
        ("Date",      BG_HDR_MAIN, "center"),
        ("Home #",    BG_HDR_HOME, "center"),
        ("Home Team", BG_HDR_HOME, "left"),
        ("Away #",    BG_HDR_AWAY, "center"),
        ("Away Team", BG_HDR_AWAY, "left"),
        ("Bar",       BG_HDR_BAR,  "left"),
    ]
    for col, (label, bg, align_h) in enumerate(hdr_defs, start=1):
        _sched_header_cell(ws, 3, col, label, bg, align_h)
    ws.row_dimensions[3].height = 20

    # ── Data rows ──────────────────────────────────────────────────────────────
    cur_row = 4
    for week_idx, (round_num, round_data) in enumerate(rounds.items()):
        row_fill    = _fill(BG_ROW_ODD if week_idx % 2 == 0 else BG_ROW_EVEN)
        group_start = cur_row

        for match_idx, match in enumerate(round_data["matches"]):
            first = (match_idx == 0)
            home  = match.home_team
            away  = match.away_team

            # Local helper to reduce repetition when writing data cells.
            def _rc(col, value, bold=False, italic=False,
                    color=FG_DARK, size=11, align_h="left"):
                cell           = ws.cell(row=cur_row, column=col, value=value)
                cell.fill      = row_fill
                cell.font      = _font(bold=bold, italic=italic, color=color, size=size)
                cell.alignment = _align(h=align_h)
                return cell

            # Week number and date only appear in the first row of each group.
            # Subsequent rows in the same round leave those cells blank.
            _rc(1, round_num if first else None, bold=True, size=10, align_h="center")
            _rc(2, round_data["date"].strftime("%b %-d") if first else None,
                size=10, align_h="center")

            # Team numbers and names. Numbers are slightly larger (12pt) to make
            # them easy to scan. Bar name is italic and muted — supporting info.
            _rc(3, home.number if home and home.number is not None else "",
                bold=True, size=12, align_h="center")
            _rc(4, home.name if home else "", size=11)
            _rc(5, away.number if away and away.number is not None else "",
                size=12, align_h="center")
            _rc(6, away.name if away else "", size=11)
            _rc(7, match.bar.name if match.bar else "",
                italic=True, color=FG_MUTED, size=10)

            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

        # Bye row — shown at the end of the round group if a team has a bye.
        # Displayed in muted italic text on a light grey background.
        if round_data.get("bye") and round_data["bye"].team:
            team     = round_data["bye"].team
            bye_fill = _fill(BG_BYE)

            # Fill all cells in the row with the bye background.
            for col in range(1, _SCHED_COLS + 1):
                ws.cell(row=cur_row, column=col).fill = bye_fill

            # "Bye:" label, team number, team name — spread across columns 3–5.
            ws.cell(row=cur_row, column=3, value="Bye:").font = _font(italic=True, color=FG_MUTED, size=10)
            ws.cell(row=cur_row, column=3).alignment = _align(h="right")
            ws.cell(row=cur_row, column=3).fill      = bye_fill

            num_val = team.number if team.number is not None else ""
            ws.cell(row=cur_row, column=4, value=num_val).font = _font(bold=True, italic=True, color=FG_MUTED, size=10)
            ws.cell(row=cur_row, column=4).alignment = _align(h="center")
            ws.cell(row=cur_row, column=4).fill      = bye_fill

            ws.cell(row=cur_row, column=5, value=team.name).font = _font(italic=True, color=FG_MUTED, size=10)
            ws.cell(row=cur_row, column=5).alignment = _align(h="left")
            ws.cell(row=cur_row, column=5).fill      = bye_fill

            ws.row_dimensions[cur_row].height = 15
            cur_row += 1

        # Apply the bold outline border around the entire week group.
        _outline_range(ws, group_start, cur_row - 1, 1, _SCHED_COLS)

    # ── Column widths ─────────────────────────────────────────────────────────
    for col, width in enumerate([5, 10, 7, 24, 7, 24, 22], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze title + header rows so they stay visible while scrolling.
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Sheet 2 — Teams
# ---------------------------------------------------------------------------

def _build_teams_sheet(ws, season):
    """
    Populate the Teams worksheet with the season's team roster.

    Columns: # | Team Name | Home Bar

    Teams are sorted by number (numbered teams first, then unnumbered teams
    alphabetically). Alternating row colours are applied for readability.
    A bold outline box wraps the entire data table.

    Args:
        ws:     The Teams worksheet.
        season: Season model instance.
    """
    sorted_teams = sorted(
        season.teams,
        key=lambda t: (t.number is None, t.number or 0, t.name)
    )
    NUM_COLS = 3

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
    c           = ws["A1"]
    c.value     = "TEAMS"
    c.fill      = _fill(BG_TITLE)
    c.font      = _font(bold=True, color=FG_WHITE, size=13)
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 26

    # Column headers
    for col, (label, align_h) in enumerate(
        [("#", "center"), ("Team Name", "left"), ("Home Bar", "left")], start=1
    ):
        c           = ws.cell(row=2, column=col, value=label)
        c.fill      = _fill(BG_TEAMS_HDR)
        c.font      = _font(bold=True, color=FG_WHITE, size=10)
        c.alignment = _align(h=align_h)
        c.border    = _border(_THICK, _THICK, _THICK, _THICK)
    ws.row_dimensions[2].height = 18

    # Data rows
    for i, team in enumerate(sorted_teams):
        row  = i + 3
        fill = _fill(BG_ROW_ODD if i % 2 == 0 else BG_ROW_EVEN)

        c           = ws.cell(row=row, column=1, value=team.number if team.number is not None else "—")
        c.fill      = fill
        c.font      = _font(bold=True, size=12)
        c.alignment = _align(h="center")

        c           = ws.cell(row=row, column=2, value=team.name)
        c.fill      = fill
        c.font      = _font(size=11)
        c.alignment = _align(h="left")

        c           = ws.cell(row=row, column=3, value=team.bar.name if team.bar else "")
        c.fill      = fill
        c.font      = _font(italic=True, color=FG_MUTED, size=10)
        c.alignment = _align(h="left")

        ws.row_dimensions[row].height = 18

    if sorted_teams:
        _outline_range(ws, 3, 2 + len(sorted_teams), 1, NUM_COLS)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Sheet 3 — Bars
# ---------------------------------------------------------------------------

def _build_bars_sheet(ws, season):
    """
    Populate the Bars worksheet with the venues used in this season.

    Columns: Bar / Venue | Tables

    Only bars that have at least one team in this season are included —
    bars that aren't involved don't need to be listed. Sorted alphabetically.

    Args:
        ws:     The Bars worksheet.
        season: Season model instance.
    """
    # Collect unique bars from the season's teams. Using a set comprehension
    # to deduplicate — multiple teams from the same bar should only list the bar once.
    bars     = sorted({team.bar for team in season.teams if team.bar}, key=lambda b: b.name)
    NUM_COLS = 2

    # Title row
    ws.merge_cells("A1:B1")
    c           = ws["A1"]
    c.value     = "BARS / VENUES"
    c.fill      = _fill(BG_TITLE)
    c.font      = _font(bold=True, color=FG_WHITE, size=13)
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 26

    # Column headers
    for col, (label, align_h) in enumerate([("Bar / Venue", "left"), ("Tables", "center")], start=1):
        c           = ws.cell(row=2, column=col, value=label)
        c.fill      = _fill(BG_TEAMS_HDR)
        c.font      = _font(bold=True, color=FG_WHITE, size=10)
        c.alignment = _align(h=align_h)
        c.border    = _border(_THICK, _THICK, _THICK, _THICK)
    ws.row_dimensions[2].height = 18

    # Data rows
    for i, bar in enumerate(bars):
        row  = i + 3
        fill = _fill(BG_ROW_ODD if i % 2 == 0 else BG_ROW_EVEN)

        c           = ws.cell(row=row, column=1, value=bar.name)
        c.fill      = fill
        c.font      = _font(size=11)
        c.alignment = _align(h="left")

        c           = ws.cell(row=row, column=2, value=bar.tables)
        c.fill      = fill
        c.font      = _font(size=11)
        c.alignment = _align(h="center")

        ws.row_dimensions[row].height = 18

    if bars:
        _outline_range(ws, 3, 2 + len(bars), 1, NUM_COLS)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def build_season_excel(season, rounds):
    """
    Build and return a complete formatted Excel workbook for a season.

    Creates a Workbook with three sheets (Schedule, Teams, Bars), delegates
    to the sheet-specific builder functions, then serialises the whole thing
    to a BytesIO buffer and returns it. The buffer is ready to be passed
    directly to Flask's send_file().

    Args:
        season: Season model instance.
        rounds: Ordered dict from _build_rounds() — {round_num: round_data}.

    Returns:
        BytesIO buffer containing the .xlsx file, seeked to position 0.
    """
    wb = Workbook()

    # The first sheet is created automatically by Workbook() — rename and use it.
    ws_sched       = wb.active
    ws_sched.title = "Schedule"
    _build_schedule_sheet(ws_sched, season, rounds)

    ws_teams       = wb.create_sheet("Teams")
    _build_teams_sheet(ws_teams, season)

    ws_bars        = wb.create_sheet("Bars")
    _build_bars_sheet(ws_bars, season)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_season_csv(season, rounds):
    """
    Build a CSV export of the season schedule and return it as a StringIO.

    Columns: Week, Date, Home #, Home Team, Away #, Away Team, Bar.
    One row per match. Bye rounds get a single row with Home Team = "BYE",
    Away Team = the bye team's display_name, other match fields blank.
    Returns the StringIO seeked to position 0, ready for send_file().

    Args:
        season: Season model instance (used for the filename — not needed here,
                but kept for API symmetry with build_season_excel).
        rounds: Ordered dict of round_num → {matches: [...], bye: Bye|None, date: date}
                as returned by _build_rounds().
    """
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(['Week', 'Date', 'Home #', 'Home Team', 'Away #', 'Away Team', 'Bar'])

    for round_num, round_data in rounds.items():
        date_str = round_data['date'].strftime('%Y-%m-%d')
        for match in round_data['matches']:
            writer.writerow([
                round_num,
                date_str,
                match.home_team.number if match.home_team.number is not None else '',
                match.home_team.name,
                match.away_team.number if match.away_team.number is not None else '',
                match.away_team.name,
                match.bar.name if match.bar else '',
            ])
        if round_data['bye']:
            writer.writerow([
                round_num,
                date_str,
                '', 'BYE',
                '', round_data['bye'].team.display_name,
                '',
            ])

    output.seek(0)
    return output
