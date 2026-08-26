"""
tests/test_export.py — Excel and CSV export content (app/main/export.py via
the /seasons/<id>/export and /seasons/<id>/export/csv routes).

These parse the actual generated files rather than stopping at status codes
— the whole point of an export route is the content, not just that *a*
response came back.
"""

import csv
import io

import openpyxl
import pytest

from app import db
from app.models import Team


EXPECTED_CSV_HEADER = ['Week', 'Date', 'Home #', 'Home Team', 'Away #', 'Away Team', 'Bar']


@pytest.fixture
def even_season_id(create_season, sample_league):
    """4 teams (sample_league), 3 weekly rounds — one full cycle, no byes."""
    return create_season('Winter Export 2026', sample_league['team_ids'], num_weeks=3)


@pytest.fixture
def odd_season_id(app, create_season, sample_league):
    """5 teams (sample_league + one extra) — exercises the bye row in exports."""
    with app.app_context():
        wildcards = Team(name='Wildcards', bar_id=sample_league['bar_ids'][1])
        db.session.add(wildcards)
        db.session.commit()
        team_ids = sample_league['team_ids'] + [wildcards.id]
    return create_season('Odd Export 2026', team_ids, num_weeks=5)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def test_excel_export_has_expected_sheets_headers_and_row_counts(admin_client, even_season_id):
    response = admin_client.get('/seasons/{0}/export'.format(even_season_id))
    assert response.status_code == 200
    assert response.content_type == (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    assert 'Winter_Export_2026_Schedule.xlsx' in response.headers['Content-Disposition']

    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    assert wb.sheetnames == ['Schedule', 'Teams', 'Bars']

    sched = wb['Schedule']
    assert sched['A1'].value == 'WINTER EXPORT 2026'
    assert [sched.cell(row=3, column=c).value for c in range(1, 8)] == [
        'Wk', 'Date', 'Home #', 'Home Team', 'Away #', 'Away Team', 'Bar']

    all_rows = list(sched.iter_rows(min_row=4, values_only=True))
    match_rows = [row for row in all_rows if row[3]]  # Home Team column populated
    assert len(match_rows) == 6  # 3 rounds x 2 matches, even teams -> no byes

    week_numbers = sorted({row[0] for row in all_rows if row[0] is not None})
    assert week_numbers == [1, 2, 3]

    teams = wb['Teams']
    assert teams['A1'].value == 'TEAMS'
    assert [teams.cell(row=2, column=c).value for c in range(1, 4)] == [
        '#', 'Team Name', 'Home Bar']
    team_names = [teams.cell(row=r, column=2).value for r in range(3, 7)]
    # No team has a league number, so the Teams-sheet sort falls back to
    # alphabetical.
    assert team_names == ['Bank Shots', 'Breakers', 'Hustlers', 'Sharks']

    bars = wb['Bars']
    assert bars['A1'].value == 'BARS / VENUES'
    bar_rows = [(bars.cell(row=r, column=1).value, bars.cell(row=r, column=2).value)
                for r in range(3, 5)]
    # Sorted by name; the Tables column is bar.tables (physical count = 4 in
    # sample_league), deliberately distinct from tables_in_use (2) so this
    # assertion would fail if the sheet accidentally showed the wrong number.
    assert bar_rows == [('Rack Room', 4), ('The Cue Club', 4)]


def test_excel_export_renders_a_bye_row_per_round_for_odd_team_count(admin_client, odd_season_id):
    response = admin_client.get('/seasons/{0}/export'.format(odd_season_id))
    assert response.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    sched = wb['Schedule']
    bye_labels = [
        sched.cell(row=r, column=3).value
        for r in range(4, sched.max_row + 1)
        if sched.cell(row=r, column=3).value == 'Bye:'
    ]
    assert len(bye_labels) == 5  # 5-team season -> exactly one bye per round


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def test_csv_export_has_expected_header_and_match_rows(admin_client, even_season_id):
    response = admin_client.get('/seasons/{0}/export/csv'.format(even_season_id))
    assert response.status_code == 200
    assert response.content_type.startswith('text/csv')
    assert 'Winter_Export_2026_Schedule.csv' in response.headers['Content-Disposition']

    rows = list(csv.reader(io.StringIO(response.data.decode('utf-8-sig'))))
    assert rows[0] == EXPECTED_CSV_HEADER
    data_rows = rows[1:]
    assert len(data_rows) == 6  # 3 rounds x 2 matches, no byes
    assert all(row[3] != 'BYE' for row in data_rows)
    assert sorted(int(row[0]) for row in data_rows) == [1, 1, 2, 2, 3, 3]


def test_csv_export_includes_bye_rows_for_odd_team_count(admin_client, odd_season_id):
    response = admin_client.get('/seasons/{0}/export/csv'.format(odd_season_id))
    assert response.status_code == 200

    rows = list(csv.reader(io.StringIO(response.data.decode('utf-8-sig'))))
    data_rows = rows[1:]
    assert len(data_rows) == 15  # 5 rounds x (2 matches + 1 bye row)

    bye_rows = [row for row in data_rows if row[3] == 'BYE']
    assert len(bye_rows) == 5  # one bye per round in a 5-team season
    for row in bye_rows:
        assert row[2] == '' and row[4] == '' and row[6] == ''  # blank match-only fields
        assert row[5]  # away column holds the bye team's display name
